import cv2
import cvzone
from cvzone.FaceMeshModule import FaceMeshDetector
from cvzone.PlotModule import LivePlot
from openpyxl import Workbook
import easygui

cap = cv2.VideoCapture(easygui.fileopenbox()) # easygui.fileopenbox() "out.mp4"
detector = FaceMeshDetector(maxFaces=1)
plotY = LivePlot(640, 360, [0, 50],  invert=True)

idList = [33, 160,158,133, 153,144, 362, 385, 387, 263, 373, 380]
ratioList = []
rawear = []
blinkstamp = []
blinkCounter = 0
counter = 0
color = (255, 0, 255)

while True:
   
    if cap.get(cv2.CAP_PROP_POS_FRAMES) == cap.get(cv2.CAP_PROP_FRAME_COUNT):
        break
    success, img = cap.read()
    img, faces = detector.findFaceMesh(img, draw=True)

    if faces:
        face = faces[0]
        # for id in idList:
        #     cv2.putText(img,str(id), face[id], cv2.FONT_HERSHEY_SIMPLEX, 0.3, color)

        lp1 = face[33]
        lp2 = face[160]
        lp3 = face[158]
        lp4 = face[133]
        lp5 = face[153]
        lp6 = face[144]
        lA, _ = detector.findDistance(lp2, lp6)
        lB, _ = detector.findDistance(lp3, lp5)
        lC, _ = detector.findDistance(lp1, lp4)
        
        lear = (lA + lB)/(2.0 * lC)

        # cv2.line(img, leftUp, leftDown, (0, 200, 0), 3)
        # cv2.line(img, leftLeft, leftRight, (0, 200, 0), 3)

        rp1 = face[362]
        rp2 = face[385]
        rp3 = face[387]
        rp4 = face[263]
        rp5 = face[373]
        rp6 = face[380]
        rA, _ = detector.findDistance(rp2, rp6)
        rB, _ = detector.findDistance(rp3, rp5)
        rC, _ = detector.findDistance(rp1, rp4)

        # cv2.line(img, leftUp, leftDown, (0, 200, 0), 3)
        # cv2.line(img, leftLeft, leftRight, (0, 200, 0), 3)

        rear = (rA + rB)/(2.0 * rC)

        ratio = (lear + rear)/2.0
        ratioList.append(ratio * 100)
        rawear.append(ratio)
        if len(ratioList) > 3:
            ratioList.pop(0)
        ratioAvg = sum(ratioList) / len(ratioList)

        if ratioAvg < 25 and counter == 0:
            blinkCounter += 1
            blinkstamp.append(cap.get(cv2.CAP_PROP_POS_FRAMES) / cap.get(cv2.CAP_PROP_FPS))
            color = (0,200,0)
            counter = 1
        if counter != 0:
            counter += 1
            if counter > 10:
                counter = 0
                color = (255,0, 255)

        cvzone.putTextRect(img, f'Blink Count: {blinkCounter}', (50, 100),
                           colorR=color)

        imgPlot = plotY.update(ratioAvg, color)
        img = cv2.resize(img, (640, 360))
        imgStack = cvzone.stackImages([img, imgPlot], 2, 1)
    else:
        img = cv2.resize(img, (640, 360))
        imgStack = cvzone.stackImages([img, img], 2, 1)

    cv2.imshow("Image", imgStack)
    key = cv2.waitKey(25)
    try:
        assert cv2.getWindowProperty("Image", 0) >= 0
    except:
        break
   

cv2.destroyAllWindows()


workbook = Workbook()
sheet = workbook.active

sheet.cell(row=1, column=1).value = "Blink Count"
sheet.cell(row=1, column=2).value = blinkCounter

for i in range(len(blinkstamp)):
    sheet.cell(row=i+2, column=1).value = blinkstamp[i]

workbook.save(filename="data.xlsx")
